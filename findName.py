import tkinter as tk
from tkinter import filedialog, simpledialog, messagebox
import re
from openpyxl import load_workbook

def main():
    # 创建主窗口
    root = tk.Tk()
    root.withdraw()  # 隐藏主窗口

    # 提示用户输入微信群成员名单
    wechat_group_members = simpledialog.askstring("输入微信群成员名单", "请粘贴微信群成员名单：")
    if not wechat_group_members:
        messagebox.showerror("错误", "未输入微信群成员名单。")
        return

    # 提示用户选择 Excel 文件
    excel_path = filedialog.askopenfilename(title="选择Excel文件", filetypes=[("Excel文件", "*.xlsx *.xls")])
    if not excel_path:
        messagebox.showerror("错误", "未选择Excel文件。")
        return

    # 处理微信群成员名单
    # 将...或者..替换成空格
    wechat_group_members = wechat_group_members.replace('...', ' ').replace('..', ' ')

    # 提取微信群成员中的中文名字
    chinese_names = re.findall(r'[\u4e00-\u9fff]+', wechat_group_members)
    wechat_group_members_list = chinese_names

    # 加载 Excel 文件
    try:
        wb = load_workbook(excel_path, read_only=True)
        ws = wb.active  # 假设数据在活动的工作表中
        excel_names = []
        for row in ws.iter_rows(min_row=2, min_col=2, max_col=2, values_only=True):  # 假设人名在第二列
            name = row[0]
            if name:
                excel_names.append(str(name).strip())
    except Exception as e:
        messagebox.showerror("错误", f"无法读取Excel文件：{e}")
        return

    # 找出不在微信群中的人（在 Excel 中但不在微信群成员列表中）
    not_in_group = [name for name in excel_names if name not in wechat_group_members_list]

    # 找出在微信群里但不在 Excel 中的人
    not_in_excel = [name for name in wechat_group_members_list if name not in excel_names]

    # 准备输出结果
    output = "以下人员未在微信群中：\n"
    output += '\n'.join(not_in_group) if not_in_group else "无"

    output += "\n\n以下人员在微信群中，但不在Excel中：\n"
    output += '\n'.join(not_in_excel) if not_in_excel else "无"

    # 在窗口中显示输出结果
    messagebox.showinfo("比较结果", output)

if __name__ == '__main__':
    main()
